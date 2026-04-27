from flask import Flask, request, send_file
from flask_cors import CORS
from openpyxl import load_workbook
import io, os, shutil, tempfile
from datetime import date

app = Flask(__name__)
CORS(app)

TEMPLATES_DIR = os.path.dirname(os.path.abspath(__file__))

TEMPLATE_FILES = {
    'A4':      'GQualité_A4_-_Prénom_Nom.xlsx',
    'A5':      'GQualité_A5_-_Prénom_Nom.xlsx',
    'A6':      'GQualité_A6_-_Prénom_Nom.xlsx',
    'A7':      'GQualité_A7_et_A8_-_Prénom_Nom.xlsx',
    'A8':      'GQualité_A7_et_A8_-_Prénom_Nom.xlsx',
    'M7':      'GQualité_M7_et_M8_-_Prénom_Nom.xlsx',
    'M8':      'GQualité_M7_et_M8_-_Prénom_Nom.xlsx',
    'M3_ACE':  'GQualité_M3_-_Prénom_Nom.xlsx',
    'M3_ECE':  'GQualité_M3_-_Prénom_Nom.xlsx',
    'M3_CVAD': 'GQualité_M3_-_Prénom_Nom.xlsx',
}

SHEET_NAMES = {
    'A4': 'A4', 'A5': 'A5  SPEKTY', 'A6': 'A6',
    'A7': 'A7', 'A8': 'A8',
    'M7': 'M7', 'M8': 'M8',
    'M3_ACE': 'M3  PAC-ACE', 'M3_ECE': 'M3  PAC-ECE', 'M3_CVAD': 'M3  CVAD',
}

OKKO_COL = {
    'A4': 2, 'A5': 2, 'A6': 2,
    'A7': 3, 'A8': 3,
    'M7': 3, 'M8': 3,
    'M3_ACE': 2, 'M3_ECE': 2, 'M3_CVAD': 2,
}

COMMENT_COL = {
    'A4': 6, 'A5': 6, 'A6': 6,
    'A7': 7, 'A8': 7,
    'M7': 7, 'M8': 7,
    'M3_ACE': 5, 'M3_ECE': 5, 'M3_CVAD': 5,
}

CRITERIA_ROWS = {
    'A4':      [4,6,7,8,9,10,11,12,13,14,15,16,17,18,19,20,21,23,24,25,26,27,28,29,30,31,32,33,34,35,36,37,38,39,40,41,42,43],
    'A5':      [4,5,7,8,10,11,12,13,15,16,18,19,20,21,23,24,25,26,28,29,30,31,33,34,35,36,38,39,40,41,43,44,45,47,48,49,51,52,54,56,57,58,59,61,62,64,65,67,68,69,70,72,73,75,76,78,79],
    'A6':      [4,5,6,7,8,9,10,11,12,13,14],
    'A7':      [4,5,6,7,8,9,10,11,12,13,14,15,16,17,18],
    'A8':      [5,6,7,8,9,10,11,12,14,15,16,17,18,19,20,21,22],
    'M7':      [4,5,6,7,8,9,10,11,12,13,14,15,16,17,18],
    'M8':      [5,6,7,8,9,10,11,12,14,15,16,17,18,19,20,21,22],
    'M3_ACE':  [6,7,8,9,11,12,13,14,15,17,18,19,20,21,23,24,25,26,27,28,29,31],
    'M3_ECE':  [5,6,7,8,9,10,11,12,13,14,15],
    'M3_CVAD': [5,6,7,8,9,10,11,12,13,14,15,16],
}

COMBINED = {
    'A7A8': ['A7', 'A8'],
    'M7M8': ['M7', 'M8'],
    'M3':   ['M3_ACE', 'M3_ECE', 'M3_CVAD'],
}

with open(os.path.join(TEMPLATES_DIR, 'index.html'), 'r', encoding='utf-8') as f:
    HTML_CONTENT = f.read()

@app.route('/')
def index():
    return HTML_CONTENT

@app.route('/generate', methods=['POST'])
def generate():
    data = request.json
    grille_code = data.get('code')
    nom = data.get('nom', 'Prénom Nom')
    date_str = data.get('date', str(date.today()))
    boa = data.get('boa', '')
    answers = data.get('answers', {})
    comments = data.get('comments', {})

    codes = COMBINED.get(grille_code, [grille_code])
    template_path = os.path.join(TEMPLATES_DIR, TEMPLATE_FILES[codes[0]])

    # Copy template to temp file to preserve all styles
    with tempfile.NamedTemporaryFile(suffix='.xlsx', delete=False) as tmp:
        tmp_path = tmp.name
    shutil.copy2(template_path, tmp_path)

    wb = load_workbook(tmp_path)

    for code in codes:
        ws = wb[SHEET_NAMES[code]]
        ws['A1'] = 'Dossier : ' + nom
        ws['A2'] = 'BOA : ' + boa
        ans = answers.get(code, {})
        cmt = comments.get(code, {})
        for i, row in enumerate(CRITERIA_ROWS[code]):
            ws.cell(row=row, column=OKKO_COL[code]).value = ans.get(str(i), 'OK')
            comment = cmt.get(str(i), '')
            if comment:
                ws.cell(row=row, column=COMMENT_COL[code]).value = comment

    buffer = io.BytesIO()
    wb.save(buffer)
    buffer.seek(0)
    os.unlink(tmp_path)

    display = {'A7A8': 'A7-A8', 'M7M8': 'M7-M8', 'M3': 'M3'}.get(grille_code, grille_code)
    filename = f'{date_str} - GQualité {display} - {nom}.xlsx'

    return send_file(buffer, as_attachment=True, download_name=filename,
                     mimetype='application/vnd.openxmlformats-officedocument.spreadsheetml.sheet')

if __name__ == '__main__':
    app.run(host='0.0.0.0', port=int(os.environ.get('PORT', 5000)))
